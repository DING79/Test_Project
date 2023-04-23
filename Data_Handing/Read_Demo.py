from openpyxl import load_workbook,Workbook

# 文件路径
path = r"C:\Users\cyberway\Desktop\111.xlsx"
print("文件路径：" + path)
# 创建对象 读取文件
lw = load_workbook(path, data_only=True)
# 获取所有sheet
sheets = lw.worksheets
# #获取当前活跃的sheet
# wb = Workbook()
# ws = wb.active
print("工作簿所有sheet：" + str(sheets))
sheet_name = input("请根据上面工作簿,输入sheet名称,查看对应信息：")
for sheet_name_value in sheets:
    var = str(sheet_name_value)
    s_value = var.split('"')
    if sheet_name == s_value[1]:
        print("当前为您查找的sheet名称是：" + sheet_name)
        sheet = lw[sheet_name]
        rows = sheet.rows
        columns = sheet.columns
        print(sheet_name + "内容如下：")
        dic = {}
        # 迭代列读取
        for col in columns:
            col_value = []
            for row in col:
                col_value.append(row.value)
            dic = {col_value[0]: col_value[1:]}
            print(str(dic))
    else:
        print("请检查是否输入正确！")
        break


# #输入想要获取的sheet名称
# sheet_name = input("请根据上面工作簿,输入sheet名称,查看对应信息：")
# # print(sheet_name)
# # 读取sheet名称为Sheet1的工作表
# sheet = lw[sheet_name]
# s = str(sheet)
# var = s.split('"')
# print("您当前想要查找的sheet名称为：" + var[1])
#读取所有行
# rows = sheet.rows
# # print(rows)
# #读取所有列
# columns = sheet.columns
# print(columns)
# print(var[1] + "内容如下：")
# 读取内容
# dic = {}
# # 迭代列读取
# for col in columns:
#     col_value = []
#     for row in col:
#         col_value.append(row.value)
#     dic = {col_value[0]: col_value[1:]}
#     print(str(dic))


# 迭代行读取
# for row in rows:
#     row_value = [col.value for col in row]
#     print(row_value)


# #迭代列读取
# for col in columns:
#     col_value = []
#     for row in col:
#         col_value.append(row.value)
#     print(col_value)


# print("========================================================================================")





